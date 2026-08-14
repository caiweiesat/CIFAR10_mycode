import sys
import numpy as np
import matplotlib.pyplot as plt

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os
import numpy as np
from matplotlib.colors import LinearSegmentedColormap
"""
gray_combination用于得到器件可模拟的灰度值，
返回值1:字典的key为灰度，value为该灰度值需要的能量密度、偏振角、绝对差(用于函数内哈希排序)、电流值
返回值2:0-255的查找表lut,[灰度值,偏振角,能量密度,与目标灰度绝对差]
"""

def gray_combination(energy_density:list[list[float]]|None=None,
                     polarized_light:list[list[float|int]]|None=None,
                     energy_benchmark:float|None=None,
                     slope:float|None=None)->tuple[dict[int,list],list]:
    """
    三种情况，只有能量密度调节，只有偏振角调节，能量密度和偏振角组合调节
    """
    if energy_density is not None and polarized_light is None:
        # print("只有能量密度调节")
        min_current=energy_density[0][0]
        max_current=energy_density[-1][0]
        # 灰度 将电流映射到0-255范围
        # 哈希结构，只存放最接近整数灰度的组合
        # 整数灰度(key):[能量密度,绝对差]
        all_gray={}
        lut=[]
        for current,energy in energy_density:
            # 灰度(浮点数)
            gray_float = (current - min_current) / (max_current - min_current) * 255
            # 灰度(整数)
            gray_int = round(gray_float)
            # 绝对差
            absolute_difference = abs(gray_float - gray_int)
            # 如果该灰度已存在，则对比浮点数和整数的绝对值，保留最接近整数的组合
            if gray_int in all_gray:
                # 当前字典绝对差
                dict_absolute_difference = all_gray[gray_int][1]
                if absolute_difference < dict_absolute_difference:
                    all_gray[gray_int] = [energy, absolute_difference,current]
            else:
                all_gray[gray_int] = [energy, absolute_difference,current]

        keys_np=np.array(list(all_gray.keys()))
        for gray in range(256):
            # 得到该灰度和每个key的距离
            dif_np = np.abs(gray - keys_np)
            # 最小值下标
            min_idx = np.argmin(dif_np)
            # 选取最短距离的key
            best_key = keys_np[min_idx]

            # 当前能量密度
            cur_energy=all_gray[best_key][0]
            # 当前电流值
            cur_current=all_gray[best_key][2]
            # 电流每灰度
            per_current = (max_current - min_current) / 255.0
            # 目标电流值
            target_current = gray * per_current + min_current

            # lut.append([best_key,cur_energy,abs(target_current-cur_current)])
            lut.append([best_key, cur_energy, abs(target_current - cur_current) / per_current])
        return all_gray,lut

    elif polarized_light is not None and energy_density is None:
        # print("只有偏振角调节")
        polarized_light.sort()
        min_current=polarized_light[0][0]
        max_current=polarized_light[-1][0]
        # 灰度 将电流映射到0-255范围
        # 哈希结构，只存放最接近整数灰度的组合
        # 整数灰度(key):[偏振角,绝对差]
        all_gray = {}
        lut=[]
        for current,angle in polarized_light:
            # 灰度(浮点数)
            gray_float = (current - min_current) / (max_current - min_current) * 255
            # 灰度(整数)
            gray_int = round(gray_float)
            # 绝对差
            absolute_difference = abs(gray_float - gray_int)
            # 如果该灰度已存在，则对比浮点数和整数的绝对值，保留最接近整数的组合
            if gray_int in all_gray:
                # 当前字典绝对差
                dict_absolute_difference = all_gray[gray_int][1]
                if absolute_difference < dict_absolute_difference:
                    all_gray[gray_int] = [angle, absolute_difference,current]
            else:
                all_gray[gray_int] = [angle, absolute_difference,current]

        keys_np = np.array(list(all_gray.keys()))
        for gray in range(256):
            # 得到该灰度和每个key的距离
            dif_np = np.abs(gray - keys_np)
            # 最小值下标
            min_idx = np.argmin(dif_np)
            # 选取最短距离的key
            best_key = keys_np[min_idx]

            # 当前偏振角
            cur_angle = all_gray[best_key][0]
            # 当前电流值
            cur_current = all_gray[best_key][2]
            # 电流每灰度
            per_current = (max_current - min_current) / 255.0
            # 目标电流值
            target_current = gray * per_current + min_current

            # lut.append([best_key, cur_angle, abs(target_current - cur_current)])
            lut.append([best_key, cur_angle, abs(target_current - cur_current)/per_current])
        return all_gray,lut

    elif energy_density is not None and polarized_light is not None\
            and energy_benchmark is not None and slope is not None:
        # print("偏振角和能量密度组合调节")
        # 能量比例 记录能量对比默认值(偏振角测试的能量密度)的比例
        # 格式为[[比例,能量密度],...]
        energy_radio = []
        for _,energy in energy_density:
            # 能量变化的比例
            radio_result = energy / energy_benchmark
            # radio的slope次方
            energy_radio.append([radio_result**slope, energy])

        # 可能的电流值 由不同的能量密度和偏振角组合而来
        # 格式为[[电流值,偏振角,能量密度],...]
        possible_current = []
        for radio, energy in energy_radio:
            for current, angle in polarized_light:
                possible_current.append([radio * current, angle, energy])
        # 排序是为了后续快速取出最大和最小值
        possible_current.sort()
        min_current = possible_current[0][0]
        max_current = possible_current[-1][0]

        # 灰度 将电流映射到0-255范围
        # 哈希结构，只存放最接近整数灰度的组合
        # 整数灰度(key):[角度,能量密度,绝对差]
        all_gray = {}
        lut=[]
        for current, radio, energy in possible_current:
            # 灰度(浮点数)
            gray_float = (current - min_current) / (max_current - min_current) * 255
            # 灰度(整数)
            gray_int = round(gray_float)
            # 绝对差
            absolute_difference = abs(gray_float - gray_int)
            # 如果该灰度已存在，则对比浮点数和整数的绝对值，保留最接近整数的组合
            if gray_int in all_gray:
                # 当前字典绝对差
                dict_absolute_difference = all_gray[gray_int][2]
                if absolute_difference < dict_absolute_difference:
                    all_gray[gray_int] = [radio, energy, absolute_difference,current]
            else:
                all_gray[gray_int] = [radio, energy, absolute_difference,current]

        keys_np = np.array(list(all_gray.keys()))
        for gray in range(256):
            # 得到该灰度和每个key的距离
            dif_np = np.abs(gray - keys_np)
            # 最小值下标
            min_idx = np.argmin(dif_np)
            # 选取最短距离的key
            best_key = keys_np[min_idx]

            # 当前偏振角
            cur_angle = all_gray[best_key][0]
            # 当前能量密度
            cur_energy = all_gray[best_key][1]
            # 当前电流值
            cur_current = all_gray[best_key][3]
            # 电流每灰度
            per_current = (max_current - min_current) / 255.0
            # 目标电流值
            target_current = gray * per_current + min_current

            # lut.append([best_key, cur_angle, cur_energy, abs(target_current - cur_current)])
            lut.append([best_key, cur_angle, cur_energy, abs(target_current - cur_current)/per_current])
        return all_gray,lut

    else:raise KeyError("参数错误")

"""
gray_apx用于给指定的数据集按照查找表变化指定值
"""
def gray_apx(x_data:np.ndarray,lut:list,color:str,apx_x_data:np.ndarray|None=None)->np.ndarray:
    # 同步类型
    if apx_x_data is None:
        apx_x_data = x_data.astype(type(lut[0])).copy()
    lut_np=np.array(lut)
    if color == "R":
        color_through = 0
    elif color == "G":
        color_through = 1
    elif color == "B":
        color_through = 2
    else:
        raise KeyError("颜色参数错误")
    # 根据查找表替换
    apx_x_data[..., color_through] = lut_np[x_data[..., color_through]]
    return apx_x_data

"""
my_list_split用于将二维list的每行的第x项组成新的一维list返回
"""
def my_list_split(lut:list,x:int)->list:
    result=[]
    for row in lut:
        result.append(row[x])
    return result


def plot_angle_grid(angle_data: np.ndarray, figsize=(8, 8), save_path=None, dpi=300):
    """
    32×32角度网格可视化
    每个格子绘制一对箭头，两支箭头夹角等于给定角度
    :param angle_data: (32,32) np数组，元素取值 0,30,60,90
    :param figsize: 画布尺寸
    :param save_path: 图片保存路径，None则不保存
    :param dpi: 保存图片分辨率
    """
    assert angle_data.shape == (32, 32), "Input array must be shape (32,32)"
    N = 32
    plt.figure(figsize=figsize)
    ax = plt.gca()
    ax.set_xlim(-0.5, N - 0.5)
    ax.set_ylim(-0.5, N - 0.5)
    ax.set_aspect("equal")
    ax.invert_yaxis()

    # 底色色块
    im = ax.imshow(angle_data, cmap="plasma", alpha=0.35, vmin=0, vmax=90)

    arrow_len = 0.38    # 箭头长度，不要超出方格
    for y in range(N):
        for x in range(N):
            theta = angle_data[y, x]
            center_x, center_y = x, y
            ang1 = np.radians(0)
            ang2 = np.radians(0 - theta)

            # 第一支箭头
            dx1 = arrow_len * np.cos(ang1)
            dy1 = arrow_len * np.sin(ang1)
            ax.arrow(center_x, center_y, dx1, dy1,
                     head_width=0.06, head_length=0.08,
                     fc="black", ec="black", lw=1)
            # 第二支箭头
            dx2 = arrow_len * np.cos(ang2)
            dy2 = arrow_len * np.sin(ang2)
            ax.arrow(center_x, center_y, dx2, dy2,
                     head_width=0.06, head_length=0.08,
                     fc="black", ec="black", lw=1)

    cbar = plt.colorbar(im, ticks=[0, 30, 60, 90])
    cbar.set_label("Angle (°)")
    plt.xticks([])
    plt.yticks([])
    plt.title("32×32 Angle Distribution (Pair Arrows)")
    plt.tight_layout()

    # 保存逻辑，移除plt.show()
    if save_path is not None:
        plt.savefig(save_path, dpi=dpi, bbox_inches="tight")

    plt.close()  # 释放画布，避免循环绘图内存溢出

def plot_energy_heatmap(energy_data: np.ndarray, figsize=(8, 8), cmap="viridis",
                        vmin=None, vmax=None, save_path=None, dpi=300):
    """
    绘制32×32能量密度热力图
    :param energy_data: (32,32) numpy数组，能量密度，单位 mW/cm²
    :param figsize: 画布尺寸
    :param cmap: 配色方案，可选 viridis, plasma, coolwarm, inferno
    :param vmin: 色标下限，None自动适配
    :param vmax: 色标上限，None自动适配
    :param save_path: 图片保存路径，None则不保存
    :param dpi: 保存分辨率
    """
    assert energy_data.shape == (32, 32), "Input array must be shape (32,32)"

    plt.figure(figsize=figsize)
    ax = plt.gca()

    im = ax.imshow(energy_data, cmap=cmap, vmin=vmin, vmax=vmax)

    cbar = plt.colorbar(im)
    cbar.set_label("Energy Density (mW/cm²)")

    plt.xticks([])
    plt.yticks([])
    plt.title("Energy Density Heatmap")
    plt.tight_layout()

    # 保存图片，移除plt.show()
    if save_path is not None:
        plt.savefig(save_path, dpi=dpi, bbox_inches="tight")

    plt.close()  # 释放画布，防止循环绘图内存泄漏


def plot_sum_3ch_heatmap(rgb_data: np.ndarray, channel: str = "sum", figsize=(8, 8), vmin=None, vmax=None, save_path=None,
                         dpi=300):
    """
    输入32×32×3数组，支持选择通道或者三通道求和绘制热力图
    配色规则：0 = 白色，数值越大颜色越红

    :param rgb_data: shape=(32,32,3) numpy数组，3通道亮度误差数据
    :param channel: 选择通道: "sum"(三通道求和) / "R" / "G" / "B"
    :param figsize: 画布大小
    :param vmin: 色标最小值，**必须设置vmin=0才能保证0为白色**
    :param vmax: 色标最大值，None自动适配数据
    :param save_path: 图片保存路径，如 "./current_error.png"；为None时不保存
    :param dpi: 保存图片分辨率
    :return out_data: 输出二维数组 (32,32)，求和或者单通道切片结果
    """
    assert rgb_data.shape == (32, 32, 3), "输入数组shape必须为 (32, 32, 3)"

    if channel.upper() == "SUM":
        out_data = np.sum(rgb_data, axis=-1)
    elif channel.upper() == "R":
        out_data = rgb_data[..., 0]
    elif channel.upper() == "G":
        out_data = rgb_data[..., 1]
    elif channel.upper() == "B":
        out_data = rgb_data[..., 2]
    else:
        raise ValueError('channel 只能是 "sum", "R", "G", "B"')

    # 自定义色板：白色(0) → 浅橙 → 深红
    color_list = [
        (1.0, 1.0, 1.0),   # 白色（最小值位置）
        (1.0, 0.75, 0.6),
        (0.9, 0.2, 0.1)    # 深红色（最大值位置）
    ]
    white_red_cmap = LinearSegmentedColormap.from_list("white_red", color_list, N=256)

    plt.figure(figsize=figsize)
    ax = plt.gca()

    im = ax.imshow(out_data, cmap=white_red_cmap, vmin=vmin, vmax=vmax)

    cbar = plt.colorbar(im)
    cbar.set_label("Brightness Error")

    plt.xticks([])
    plt.yticks([])

    title_map = {
        "SUM": "Sum of Three‑Channel Brightness Error",
        "R": "R‑Channel Brightness Error",
        "G": "G‑Channel Brightness Error",
        "B": "B‑Channel Brightness Error"
    }
    plt.title(title_map[channel.upper()])
    plt.tight_layout()

    if save_path is not None:
        plt.savefig(save_path, dpi=dpi, bbox_inches="tight")

    plt.close()
    return out_data


def write_gray_data_to_excel(data_dict, file_path, sheet_name=None, tuple_mode="只有能量密度"):
    """
    将灰度‑参数数据写入Excel，每次调用新建一个分表，保持dict原始顺序。
    写入列：灰度、偏振角、能量密度、电流值（不再写入绝对差）。
    支持元组长度3 / 4，绝对差仅用于拆包定位，不输出。

    参数:
        data_dict: dict，key=灰度值
                   value元组：
                     4元素: (偏振角, 能量密度, 绝对差, 电流值)
                     3元素由tuple_mode决定：
                        "只有能量密度" → (能量密度, 绝对差, 电流值)
                        "只有偏振角"   → (偏振角, 绝对差, 电流值)
        file_path: Excel文件路径
        sheet_name: 新建工作表名称，不传自动生成
        tuple_mode: "只有能量密度" / "只有偏振角"，仅对3元素元组生效
    """
    rows = []
    for gray, vals in data_dict.items():
        if len(vals) == 4:
            # 你的元组顺序：偏振角, 能量密度, 绝对差, 电流值
            angle, energy, _abs_diff, current = vals
        elif len(vals) == 3:
            if tuple_mode == "只有能量密度":
                energy, _abs_diff, current = vals
                angle = None
            elif tuple_mode == "只有偏振角":
                angle, _abs_diff, current = vals
                energy = None
            else:
                raise ValueError('tuple_mode只能为 "只有能量密度" 或 "只有偏振角"')
        else:
            raise ValueError(f"元组长度必须为3或4，得到 {len(vals)}")

        rows.append((gray, angle, energy, current))

    # 创建/加载工作簿
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    # 处理sheet名称，避免重名
    if sheet_name is None:
        sheet_name = f"Sheet_{len(wb.sheetnames) + 1}"
    sheet_name = sheet_name[:31]
    if sheet_name in wb.sheetnames:
        sheet_name = f"{sheet_name}_{len(wb.sheetnames) + 1}"

    ws = wb.create_sheet(title=sheet_name)

    # 表头
    headers = ["灰度", "偏振角(°)", "能量密度(mW/cm²)", "电流值(nA)"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 写入数据，None会输出为空单元格
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # 自适应列宽
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 10)

    wb.save(file_path)
    return sheet_name


def write_lut_to_excel(lut, file_path, sheet_name=None, lut_mode="两者都有"):
    """
    将256长度的LUT查找表写入Excel，每次调用新建一个分表。

    参数:
        lut: list，长度256，索引0~255即目标灰度值
             每个元素为list：
               4元素: [实际灰度值, 偏振角, 能量密度, 绝对差]   → lut_mode="两者都有"
               3元素由lut_mode决定：
                 "只有偏振角" → [实际灰度值, 偏振角, 绝对差]，能量密度留空
                 "只有能量密度" → [实际灰度值, 能量密度, 绝对差]，偏振角留空
        file_path: Excel文件路径
        sheet_name: 新建工作表名称，不传自动生成
        lut_mode: "两者都有" / "只有偏振角" / "只有能量密度"

    写入列：目标灰度值、实际灰度值、偏振角(°)、能量密度(mW/cm²)
    绝对差仅用于拆包定位，不写入Excel。
    """
    rows = []
    for target_gray, item in enumerate(lut):
        if len(item) == 4:
            actual_gray, angle, energy, _abs_diff = item
        elif len(item) == 3:
            if lut_mode == "只有偏振角":
                actual_gray, angle, _abs_diff = item
                energy = None
            elif lut_mode == "只有能量密度":
                actual_gray, energy, _abs_diff = item
                angle = None
            else:
                raise ValueError('3元素LUT时，lut_mode必须为 "只有偏振角" 或 "只有能量密度"')
        else:
            raise ValueError(f"LUT元素长度必须为3或4，得到 {len(item)}")

        rows.append((target_gray, actual_gray, angle, energy))

    # 创建/加载工作簿
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    # 处理sheet名称，避免重名
    if sheet_name is None:
        sheet_name = f"Sheet_{len(wb.sheetnames) + 1}"
    sheet_name = sheet_name[:31]
    if sheet_name in wb.sheetnames:
        sheet_name = f"{sheet_name}_{len(wb.sheetnames) + 1}"

    ws = wb.create_sheet(title=sheet_name)

    # 表头（直接用 Unicode 上标 ²）
    headers = ["目标灰度值", "实际灰度值", "偏振角(°)", "能量密度(mW/cm²)"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # 写入数据
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # 自适应列宽
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(file_path)
    return sheet_name


if __name__ == "__main__":
    import torchvision.transforms as transforms
    import torch
    from pathlib import Path
    from device_test_data import *
    from read import *
    # 加载数据集
    (x_train, y_train), (x_test, y_test) = load_all_cifar10(r"D:\PythonProject\cifar10\cifar-10-batches-py")

    # 读取类别名称
    meta_path = r"D:\PythonProject\cifar10\cifar-10-batches-py\batches.meta"
    meta_data = unpickle(meta_path)
    class_names = meta_data["label_names"]
    print(type(class_names))
    """
    一共6组实验
    RGB 能量密度+偏振角
    RGB 能量密度
    RGB 偏振角
    R   能量密度+偏振角
    G   能量密度+偏振角
    B   能量密度+偏振角
    """

    # 先获取R、G、B的单能量、单偏振角、能量和偏振角组合的可调灰度值
    # 只有能量密度
    R_energy_gray,R_energy_lut = gray_combination(energy_density=R_energy_density)
    G_energy_gray,G_energy_lut = gray_combination(energy_density=G_energy_density)
    B_energy_gray,B_energy_lut = gray_combination(energy_density=B_energy_density)
    # 只有偏振角
    R_polarized_gray,R_polarized_lut = gray_combination(polarized_light=R_polarized_light)
    G_polarized_gray,G_polarized_lut = gray_combination(polarized_light=G_polarized_light)
    B_polarized_gray,B_polarized_lut = gray_combination(polarized_light=B_polarized_light)
    # 能量密度和偏振角组合
    R_energy_and_polarized_gray,R_energy_and_polarized_lut = gray_combination(energy_density=R_energy_density,
                                                   polarized_light=R_polarized_light,
                                                   energy_benchmark=R_energy_benchmark,
                                                   slope=R_slope)
    G_energy_and_polarized_gray,G_energy_and_polarized_lut = gray_combination(energy_density=G_energy_density,
                                                   polarized_light=G_polarized_light,
                                                   energy_benchmark=G_energy_benchmark,
                                                   slope=G_slope)
    B_energy_and_polarized_gray,B_energy_and_polarized_lut = gray_combination(energy_density=B_energy_density,
                                                   polarized_light=B_polarized_light,
                                                   energy_benchmark=B_energy_benchmark,
                                                   slope=B_slope)

    # 然后获取RGB三种情况下的数据集，单通道的数据集由RGB通道数据集做切片获得
    # RGB 能量密度
    R_energy_graylut = my_list_split(R_energy_lut, 0)
    G_energy_graylut = my_list_split(G_energy_lut, 0)
    B_energy_graylut = my_list_split(B_energy_lut, 0)

    RGB_energy_x_train = gray_apx(x_train, R_energy_graylut, color="R")
    RGB_energy_x_train = gray_apx(x_train, G_energy_graylut, color="G",apx_x_data=RGB_energy_x_train)
    RGB_energy_x_train = gray_apx(x_train, B_energy_graylut, color="B",apx_x_data=RGB_energy_x_train)

    RGB_energy_x_test = gray_apx(x_test, R_energy_graylut, color="R")
    RGB_energy_x_test = gray_apx(x_test, G_energy_graylut, color="G",apx_x_data=RGB_energy_x_test)
    RGB_energy_x_test = gray_apx(x_test, B_energy_graylut, color="B",apx_x_data=RGB_energy_x_test)



    # RGB 偏振角
    R_polarized_graylut = my_list_split(R_polarized_lut, 0)
    G_polarized_graylut = my_list_split(G_polarized_lut, 0)
    B_polarized_graylut = my_list_split(B_polarized_lut, 0)

    RGB_polarized_x_train = gray_apx(x_train, R_polarized_graylut, color="R")
    RGB_polarized_x_train = gray_apx(x_train, G_polarized_graylut, color="G",apx_x_data=RGB_polarized_x_train)
    RGB_polarized_x_train = gray_apx(x_train, B_polarized_graylut, color="B",apx_x_data=RGB_polarized_x_train)

    RGB_polarized_x_test = gray_apx(x_test, R_polarized_graylut, color="R")
    RGB_polarized_x_test = gray_apx(x_test, G_polarized_graylut, color="G",apx_x_data=RGB_polarized_x_test)
    RGB_polarized_x_test = gray_apx(x_test, B_polarized_graylut, color="B",apx_x_data=RGB_polarized_x_test)



    # RGB 能量密度+偏振角
    R_energy_and_polarized_graylut = my_list_split(R_energy_and_polarized_lut,0)
    G_energy_and_polarized_graylut = my_list_split(G_energy_and_polarized_lut, 0)
    B_energy_and_polarized_graylut = my_list_split(B_energy_and_polarized_lut, 0)

    RGB_energy_and_polarized_x_train = gray_apx(x_train, R_energy_and_polarized_graylut,
                                                color="R")
    RGB_energy_and_polarized_x_train = gray_apx(x_train, G_energy_and_polarized_graylut,
                                                color="G",apx_x_data=RGB_energy_and_polarized_x_train)
    RGB_energy_and_polarized_x_train = gray_apx(x_train, B_energy_and_polarized_graylut,
                                                color="B",apx_x_data=RGB_energy_and_polarized_x_train)

    RGB_energy_and_polarized_x_test = gray_apx(x_test, R_energy_and_polarized_graylut
                                               , color="R")
    RGB_energy_and_polarized_x_test = gray_apx(x_test, G_energy_and_polarized_graylut
                                               , color="G",apx_x_data=RGB_energy_and_polarized_x_test)
    RGB_energy_and_polarized_x_test = gray_apx(x_test, B_energy_and_polarized_graylut
                                               , color="B",apx_x_data=RGB_energy_and_polarized_x_test)


    """偏振角"""
    # 模拟(能量+角度)需要的偏振角
    R_energy_and_polarized_anglelut = my_list_split(R_energy_and_polarized_lut, 1)
    G_energy_and_polarized_anglelut = my_list_split(G_energy_and_polarized_lut, 1)
    B_energy_and_polarized_anglelut = my_list_split(B_energy_and_polarized_lut, 1)

    RGB_energy_and_polarized_angle = gray_apx(x_train, R_energy_and_polarized_anglelut,
                                              color="R")
    RGB_energy_and_polarized_angle = gray_apx(x_train, G_energy_and_polarized_anglelut,
                                              color="G", apx_x_data=RGB_energy_and_polarized_angle)
    RGB_energy_and_polarized_angle = gray_apx(x_train, B_energy_and_polarized_anglelut,
                                              color="B", apx_x_data=RGB_energy_and_polarized_angle)
    # 模拟(角度)需要的偏振角
    R_polarized_anglelut = my_list_split(R_polarized_lut, 1)
    G_polarized_anglelut = my_list_split(G_polarized_lut, 1)
    B_polarized_anglelut = my_list_split(B_polarized_lut, 1)

    RGB_polarized_angle = gray_apx(x_train, R_polarized_anglelut,
                                              color="R")
    RGB_polarized_angle = gray_apx(x_train, G_polarized_anglelut,
                                              color="G", apx_x_data=RGB_polarized_angle)
    RGB_polarized_angle = gray_apx(x_train, B_polarized_anglelut,
                                              color="B", apx_x_data=RGB_polarized_angle)
    """能量密度"""
    # 模拟(能量+角度)需要的能量密度
    R_energy_and_polarized_energylut = my_list_split(R_energy_and_polarized_lut, 2)
    G_energy_and_polarized_energylut = my_list_split(G_energy_and_polarized_lut, 2)
    B_energy_and_polarized_energylut = my_list_split(B_energy_and_polarized_lut, 2)

    RGB_energy_and_polarized_energy = gray_apx(x_train, R_energy_and_polarized_energylut,
                                               color="R")
    RGB_energy_and_polarized_energy = gray_apx(x_train, G_energy_and_polarized_energylut,
                                               color="G", apx_x_data=RGB_energy_and_polarized_energy)
    RGB_energy_and_polarized_energy = gray_apx(x_train, B_energy_and_polarized_energylut,
                                               color="B", apx_x_data=RGB_energy_and_polarized_energy)
    # 模拟(能量)需要的能量密度
    R_energy_energylut = my_list_split(R_energy_lut, 1)
    G_energy_energylut = my_list_split(G_energy_lut, 1)
    B_energy_energylut = my_list_split(B_energy_lut, 1)

    RGB_energy_energy = gray_apx(x_train, R_energy_energylut,
                                               color="R")
    RGB_energy_energy = gray_apx(x_train, G_energy_energylut,
                                               color="G", apx_x_data=RGB_energy_energy)
    RGB_energy_energy = gray_apx(x_train, B_energy_energylut,
                                               color="B", apx_x_data=RGB_energy_energy)
    """绝对亮度差"""
    # 模拟(能量+角度)结果的亮度绝对差
    R_energy_and_polarized_brightnesslut = my_list_split(R_energy_and_polarized_lut, 3)
    G_energy_and_polarized_brightnesslut = my_list_split(G_energy_and_polarized_lut, 3)
    B_energy_and_polarized_brightnesslut = my_list_split(B_energy_and_polarized_lut, 3)

    RGB_energy_and_polarized_current = gray_apx(x_train, R_energy_and_polarized_brightnesslut,
                                                color="R")
    RGB_energy_and_polarized_current = gray_apx(x_train, G_energy_and_polarized_brightnesslut,
                                                color="G", apx_x_data=RGB_energy_and_polarized_current)
    RGB_energy_and_polarized_current = gray_apx(x_train, B_energy_and_polarized_brightnesslut,
                                                color="B", apx_x_data=RGB_energy_and_polarized_current)
    # 模拟(能量)结果的亮度绝对差
    R_energy_brightnesslut = my_list_split(R_energy_lut, 2)
    G_energy_brightnesslut = my_list_split(G_energy_lut, 2)
    B_energy_brightnesslut = my_list_split(B_energy_lut, 2)

    RGB_energy_current = gray_apx(x_train, R_energy_brightnesslut,
                                                color="R")
    RGB_energy_current = gray_apx(x_train, G_energy_brightnesslut,
                                                color="G", apx_x_data=RGB_energy_current)
    RGB_energy_current = gray_apx(x_train, B_energy_brightnesslut,
                                                color="B", apx_x_data=RGB_energy_current)
    # 模拟(角度)结果的亮度绝对差
    R_polarized_brightnesslut = my_list_split(R_polarized_lut, 2)
    G_polarized_brightnesslut = my_list_split(G_polarized_lut, 2)
    B_polarized_brightnesslut = my_list_split(B_polarized_lut, 2)

    RGB_polarized_current = gray_apx(x_train, R_polarized_brightnesslut,
                                                color="R")
    RGB_polarized_current = gray_apx(x_train, G_polarized_brightnesslut,
                                                color="G", apx_x_data=RGB_polarized_current)
    RGB_polarized_current = gray_apx(x_train, B_polarized_brightnesslut,
                                                color="B", apx_x_data=RGB_polarized_current)
    """分割线"""
    my_tasks=("energy_and_polarized","energy","polarized")
    color_list=["R","G","B"]
    pic_nums=[0,2,7]
    for task in my_tasks:
        if task == "energy_and_polarized":
            RGB_angle = RGB_energy_and_polarized_angle
            RGB_energy = RGB_energy_and_polarized_energy
            RGB_current = RGB_energy_and_polarized_current
            RGB_x_train = RGB_energy_and_polarized_x_train
        elif task == "energy":
            RGB_angle = None
            RGB_energy = RGB_energy_energy
            RGB_current = RGB_energy_current
            RGB_x_train = RGB_energy_x_train
        elif task == "polarized":
            RGB_angle = RGB_polarized_angle
            RGB_energy = None
            RGB_current = RGB_polarized_current
            RGB_x_train = RGB_polarized_x_train
        else:
            RGB_angle = None
            RGB_energy = None
            RGB_current = None
            RGB_x_train = None
        for pic_num in pic_nums:
            my_p=pic_num
            for through,color in enumerate(color_list):
                # 偏振角
                if RGB_angle is not None:
                    test_arr=RGB_angle[my_p,:,:,through]
                    target_dir = f"./DemoPic/{task}/picture{pic_num}"
                    Path(target_dir).mkdir(parents=True, exist_ok=True)
                    # 先转为Path对象
                    target_dir = Path(target_dir)
                    # 拼接文件完整路径
                    save_file = target_dir / f"{color}_angle_map.png"
                    plot_angle_grid(test_arr, save_path=save_file)

                # 能量密度
                if RGB_energy is not None:
                    test_energy=RGB_energy[my_p,:,:,through]
                    target_dir = f"./DemoPic/{task}/picture{pic_num}"
                    Path(target_dir).mkdir(parents=True, exist_ok=True)
                    # 先转为Path对象
                    target_dir = Path(target_dir)
                    # 拼接文件完整路径
                    save_file = target_dir / f"{color}_energy_map.png"
                    plot_energy_heatmap(test_energy, save_path=save_file)

            # 亮度绝对差
            test_3ch=RGB_current[my_p,:,:,:]
            target_dir = f"./DemoPic/{task}/picture{pic_num}"
            Path(target_dir).mkdir(parents=True, exist_ok=True)
            # 先转为Path对象
            target_dir = Path(target_dir)
            # 拼接文件完整路径
            save_file = target_dir / f"brightness_error_sum.png"
            plot_sum_3ch_heatmap(test_3ch, vmin=0,vmax=200, save_path=save_file)

            plot_sum_3ch_heatmap(test_3ch, channel="R", vmin=0, vmax=80,
                                 save_path=target_dir / f"brightness_error_R.png")
            plot_sum_3ch_heatmap(test_3ch, channel="G", vmin=0, vmax=80,
                                 save_path=target_dir / f"brightness_error_G.png")
            plot_sum_3ch_heatmap(test_3ch, channel="B", vmin=0, vmax=80,
                                 save_path=target_dir / f"brightness_error_B.png")

            # 变化前
            plt.imshow(x_train[my_p])
            target_dir = f"./DemoPic/{task}/picture{pic_num}"
            Path(target_dir).mkdir(parents=True, exist_ok=True)
            # 先转为Path对象
            target_dir = Path(target_dir)
            # 拼接文件完整路径
            save_file = target_dir / "before_change.png"
            plt.savefig(save_file, dpi=300, bbox_inches="tight")
            plt.close()

            # 变化后
            plt.imshow(RGB_x_train[my_p])
            target_dir = f"./DemoPic/{task}/picture{pic_num}"
            Path(target_dir).mkdir(parents=True, exist_ok=True)
            # 先转为Path对象
            target_dir = Path(target_dir)
            # 拼接文件完整路径
            save_file = target_dir / "after_change.png"
            plt.savefig(save_file, dpi=300, bbox_inches="tight")
            plt.close()